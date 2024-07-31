# Importing the libraries used in this project
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep

# Importing the sheet's useds
workbook_path = 'Files-9Website-Info-Extraction/Lista de Ações.xlsx'
workbook = openpyxl.load_workbook(workbook_path)
search_page = workbook['Pesquisa']
final_page = workbook['Resultado']

# Clear the final page starting from the second row
final_page.delete_rows(1, final_page.max_row)

# Add headers to the final_page
headers = [
    'ATIVO', 'TIPO', 'COTAÇÃO', 
    'Rent. 1 mês', 'Rent. 3 meses', 'Rent. 1 ano', 'Rent. 2 anos', 'Rent. 5 anos', 'Rent. 10 anos',
    'Rent. Real 3 meses', 'Rent. Real 1 ano', 'Rent. Real 2 anos', 'Rent. Real 5 anos', 'Rent. Real 10 anos']
final_page.append(headers)

# Check if there are data in the search page
if search_page.max_row < 2:
    print("No data to process in the search page.")
else:
    # Opening the website
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 15)

    # Creating the search loop
    for row in search_page.iter_rows(min_row=2, values_only=True):
        ATIVO, TIPO = row
        # Open the website with the complete link
        url_path = 'https://investidor10.com.br/'
        complete_url = f"{url_path}{TIPO}{ATIVO}"  # Ensure the URL is properly formatted
        driver.get(complete_url)

        try:
            # Try to find the element and extract the text, processing it directly
            cotação_clean = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='_card-body']//span[@class='value']"))).text.replace('R$ ', '').strip()

            # Extract the values for each month and other periods
            values = []
            for i in range(2, 8):  # Extracting the previous periods
                xpath = f"//*[@id='ticker']/section/div/div[2]/div/div/div[{i}]/span"
                try:
                    value = wait.until(EC.presence_of_element_located((By.XPATH, xpath))).text
                except Exception as e:
                    value = None  # Handle cases where the element might not be found
                values.append(value)

            # Extract the additional values for Rent. Real
            for i in range(10, 15):  # Extracting Rent. Real periods
                xpath = f"//*[@id='ticker']/section/div/div[2]/div/div/div[{i}]/span"
                try:
                    value = wait.until(EC.presence_of_element_located((By.XPATH, xpath))).text
                except Exception as e:
                    value = None  # Handle cases where the element might not be found
                values.append(value)

            # Append data to the final page
            final_page.append([ATIVO, TIPO, cotação_clean] + values)

        except Exception as e:
            # Capture and print the error message
            print(f"Error for {ATIVO} - {TIPO}: {e}")

        # Sleep to avoid overwhelming the server
        sleep(1)

    # Save the workbook after all rows have been processed
    workbook.save(workbook_path)

    # Close the browser
    driver.quit()
